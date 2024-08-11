import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta

def find_latest_sales_update(directory):
    sales_update_files = list(directory.glob("*Sales Update*.xlsx"))
    if not sales_update_files:
        raise FileNotFoundError("No Sales Update file found.")
    return max(sales_update_files, key=os.path.getctime)

def convert_to_numeric(value):
    try:
        return pd.to_numeric(value, errors='coerce')
    except:
        return pd.np.nan

def update_scoreboard(import_path):
    excel_file_path = find_latest_sales_update(import_path)
    print(f"Updating scoreboard in file: {excel_file_path.name}")

    # Load the workbook without data_only to preserve formulas
    book = load_workbook(excel_file_path, data_only=False)
    book.calculation.calcMode = 'auto'
    book.save(excel_file_path)
    # Process '2023 Estimate Summary' sheet
    estimate_sheet = book['2023 Estimate Summary']
    estimate_data = [[cell.value for cell in row[:15]] for row in estimate_sheet.iter_rows()]
    estimate_columns = estimate_data[0]
    estimate_df = pd.DataFrame(estimate_data, columns=estimate_columns)

    print(f"2023 Estimate Summary processed: {len(estimate_df)} rows")
    

    # Convert CustomerNumber to numeric
    estimate_df['CustomerNumber'] = estimate_df['CustomerNumber'].apply(convert_to_numeric)
    estimate_df = estimate_df.dropna(subset=['CustomerNumber'])
    print(f"Rows after converting CustomerNumber to numeric: {len(estimate_df)}")

    estimate_df['ResidentialOrCommercial'] = estimate_df['ResidentialOrCommercial'].fillna('').astype(str)
    # print("Converted ResidentialOrCommercial to string")
    
    # print(estimate_df.dtypes)
    # print(estimate_df['ResidentialOrCommercial'].head())

    residential_df = estimate_df[estimate_df['ResidentialOrCommercial'].fillna('').astype(str).str.strip().str.upper() == 'R']
    print(f'Printing residential_df head:\n{residential_df.head()}')
    # print(f"Residential customers: {len(residential_df)}")
    # print(f'Residential df head{residential_df.head()}')
        
    # Load the 2022_pbt CSV file
    pbt_file_path = import_path / '2022_pbt.csv'
    if not pbt_file_path.exists():
        raise FileNotFoundError("2022_pbt.csv file not found in the specified directory.")
    
    production_df = pd.read_csv(pbt_file_path, low_memory=False)
    
    # Convert CustomerNumber in production_df to numeric
    production_df['CustomerNumber'] = production_df['CustomerNumber'].apply(convert_to_numeric)
    production_df = production_df.dropna(subset=['CustomerNumber'])
    production_customers = set(production_df['CustomerNumber'])
    print(f"Customers in 2022 production report: {len(production_customers)}")

    # Step 2: Filter for customers not in 2023 production report
    filtered_df = residential_df[~residential_df['CustomerNumber'].isin(production_customers)]
    print(f"Residential customers not in 2022 production report: {len(filtered_df)}")
    print(f'printing filtered_df:\n{filtered_df.head()}')

    # Step 3: Filter for EstimateRequestDate less than the value in cell R1
    
    cutoff_date = datetime.now().date() - timedelta(days=367)
    
    print(f"Using cutoff date: {cutoff_date}")

    
    filtered_df['EstimateRequestedDate'] = pd.to_datetime(filtered_df['EstimateRequestedDate'], format='%m-%d-%Y', errors='coerce').dt.date
    filtered_df = filtered_df[filtered_df['EstimateRequestedDate'] <= cutoff_date]
    print(f"Customers after filtering by EstimateRequestedDate: {len(filtered_df)}")

    # Step 4: Get unique customer numbers
    unique_customers = filtered_df['CustomerNumber'].unique()
    print(f"Unique customer numbers after all filtering: {len(unique_customers)}")

    # Update 'Scoreboard' sheet
    scoreboard_sheet = book['Scoreboard']
    # Clear contents of ONLY column O starting from row 4
    for row in range(4, scoreboard_sheet.max_row + 1):
        scoreboard_sheet.cell(row=row, column=15).value = None

    print(f"Cleared existing data in Scoreboard column O from row 4 to {scoreboard_sheet.max_row}")

    # Add unique customer numbers to column O starting from row 4
    for i, customer in enumerate(unique_customers, start=4):
        scoreboard_sheet.cell(row=i, column=15, value=int(customer))

    # Save the workbook
    book.save(excel_file_path)
    
    print(f"Scoreboard updated with {len(unique_customers)} unique customers.")
    print("Workbook saved with formulas preserved.")

if __name__ == "__main__":
    import_path = Path("C:/Users/Shadow/projects/analyitics_automator")
    update_scoreboard(import_path)