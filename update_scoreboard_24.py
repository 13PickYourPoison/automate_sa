import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import os

def find_latest_sales_update(directory):
    sales_update_files = list(directory.glob("*Sales Update*.xlsx"))
    if not sales_update_files:
        raise FileNotFoundError("No Sales Update file found.")
    return max(sales_update_files, key=os.path.getctime)

def convert_to_numeric(value):
    try:
        return pd.to_numeric(value, errors='coerce')
    except:
        return pd.np.nan

def update_scoreboard(import_path):
    excel_file_path = find_latest_sales_update(import_path)
    print(f"Updating scoreboard in file: {excel_file_path.name}")

    # Load the workbook without data_only to preserve formulas
    book = load_workbook(excel_file_path, data_only=False)

    # Process '2024 Estimate Summary' sheet
    estimate_sheet = book['2024 Estimate Summary']
    estimate_data = [[cell.value for cell in row] for row in estimate_sheet.iter_rows()]
    estimate_columns = estimate_data[0]
    estimate_df = pd.DataFrame(estimate_data[1:], columns=estimate_columns)

    print(f"Total rows in 2024 Estimate Summary: {len(estimate_df)}")

    # Convert CustomerNumber to numeric
    estimate_df['CustomerNumber'] = estimate_df['CustomerNumber'].apply(convert_to_numeric)
    estimate_df = estimate_df.dropna(subset=['CustomerNumber'])
    print(f"Rows after converting CustomerNumber to numeric: {len(estimate_df)}")

    # Step 1: Filter for Residential customers
    residential_df = estimate_df[estimate_df['ResidentialOrCommercial'].astype(str).str.strip().str.upper() == 'R']
    print(f"Residential customers: {len(residential_df)}")

    # Load the 2023_pbt CSV file
    pbt_file_path = import_path / '2023_pbt.csv'
    if not pbt_file_path.exists():
        raise FileNotFoundError("2023_pbt.csv file not found in the specified directory.")
    
    production_df = pd.read_csv(pbt_file_path, low_memory=False)
    
    if 'CustomerNumber' not in production_df.columns:
        raise ValueError("'CustomerNumber' column not found in 2023_pbt.csv")
    
    # Convert CustomerNumber in production_df to numeric
    production_df['CustomerNumber'] = production_df['CustomerNumber'].apply(convert_to_numeric)
    production_df = production_df.dropna(subset=['CustomerNumber'])
    production_customers = set(production_df['CustomerNumber'])
    print(f"Customers in 2023 production report: {len(production_customers)}")

    # Step 2: Filter for customers not in 2023 production report
    filtered_df = residential_df[~residential_df['CustomerNumber'].isin(production_customers)]
    print(f"Residential customers not in 2023 production report: {len(filtered_df)}")

    # Step 3: Get unique customer numbers
    unique_customers = filtered_df['CustomerNumber'].unique()
    print(f"Unique customer numbers after filtering: {len(unique_customers)}")

    # Update 'Scoreboard' sheet
    scoreboard_sheet = book['Scoreboard']

    # Clear contents of ONLY column C starting from row 4
    for row in range(4, scoreboard_sheet.max_row + 1):
        scoreboard_sheet.cell(row=row, column=3).value = None

    print(f"Cleared existing data in Scoreboard column C from row 4 to {scoreboard_sheet.max_row}")

    # Add unique customer numbers to column C starting from row 4
    for i, customer in enumerate(unique_customers, start=4):
        scoreboard_sheet.cell(row=i, column=3, value=int(customer))

    # Save the workbook
    book.save(excel_file_path)
    
    print(f"Scoreboard updated with {len(unique_customers)} unique customers.")
    print("Workbook saved with formulas preserved.")

if __name__ == "__main__":
    import_path = Path("C:/Users/Shadow/projects/analyitics_automator")
    update_scoreboard(import_path)