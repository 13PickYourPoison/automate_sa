import pandas as pd
from pathlib import Path
import os
from openpyxl import load_workbook

def find_latest_file(directory, pattern):
    matching_files = list(directory.glob(f"*{pattern}*"))
    if not matching_files:
        raise FileNotFoundError(f"No file matching '{pattern}' found.")
    return max(matching_files, key=os.path.getctime)

def convert_to_numeric(value):
    try:
        return pd.to_numeric(value, errors='coerce')
    except:
        return pd.np.nan

def process_csv(file_path, columns, filter_zero=None):
    df = pd.read_csv(file_path, usecols=columns, low_memory=False)
    for col in columns:
        if col in ['CustomerNumber', 'GrossSalesAmount', 'TotalPrice']:
            df[col] = df[col].apply(convert_to_numeric)
    df = df.dropna(subset=[col for col in columns if col != 'BranchNumberOfCustomer'])
    
    if filter_zero:
        df = df[df[filter_zero] != 0]
    
    return df

def update_sales_report(import_path):
    # Process Updated Sales Report CSV
    sales_report_csv = find_latest_file(import_path, "Updated_Sales_Report")
    sales_report_df = process_csv(sales_report_csv, ['CustomerNumber', 'BranchNumberOfCustomer', 'TotalPrice'], filter_zero='TotalPrice')
    print(f"Processed {sales_report_csv.name}: {len(sales_report_df)} rows")

    # Process Production By Technician CSV
    production_csv = find_latest_file(import_path, "Production_By_Technician")
    production_df = process_csv(production_csv, ['CustomerNumber', 'BranchNumberOfCustomer', 'GrossSalesAmount'], filter_zero='GrossSalesAmount')
    print(f"Processed {production_csv.name}: {len(production_df)} rows")

    # Find and update the latest Sales Update Excel file
    excel_file_path = find_latest_file(import_path, "Sales Update")
    print(f"Updating Excel file: {excel_file_path.name}")

    # Load the workbook
    book = load_workbook(excel_file_path, data_only=False)

    # Select the "Updated Sales Report" sheet
    sheet_name = "Updated Sales Report"
    if sheet_name not in book.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
    sheet = book[sheet_name]

    # Update columns C, D, E with sales_report_df data
    for col, df_col in zip(['C', 'D', 'E'], sales_report_df.columns):
        for cell in sheet[f'{col}5:{col}{sheet.max_row}']:
            cell[0].value = None
        for i, value in enumerate(sales_report_df[df_col], start=5):
            sheet.cell(row=i, column=ord(col)-64, value=value)

    # Update columns H, I, J with production_df data
    for col, df_col in zip(['H', 'I', 'J'], production_df.columns):
        for cell in sheet[f'{col}5:{col}{sheet.max_row}']:
            cell[0].value = None
        for i, value in enumerate(production_df[df_col], start=5):
            sheet.cell(row=i, column=ord(col)-64, value=value)

    # Create lists of unique CustomerNumbers
    unique_customers_C = set(sales_report_df['CustomerNumber'])
    unique_customers_H = set(production_df['CustomerNumber'])
    unique_customers_H_not_in_C = unique_customers_H - unique_customers_C

    # Update "Customer Count" sheet
    customer_count_sheet = book["Customer Count"]

    # Clear and update column C
    for cell in customer_count_sheet['C4:C' + str(customer_count_sheet.max_row)]:
        cell[0].value = None
    for i, customer in enumerate(unique_customers_C, start=4):
        customer_count_sheet.cell(row=i, column=3, value=customer)

    # Clear and update column H
    for cell in customer_count_sheet['H4:H' + str(customer_count_sheet.max_row)]:
        cell[0].value = None
    for i, customer in enumerate(unique_customers_H_not_in_C, start=4):
        customer_count_sheet.cell(row=i, column=8, value=customer)

    # Save the workbook
    book.save(excel_file_path)

    print(f"Updated sales report data: {len(sales_report_df)} rows")
    print(f"Updated production data: {len(production_df)} rows")
    print(f"Updated Customer Count sheet: {len(unique_customers_C)} customers in column C, {len(unique_customers_H_not_in_C)} customers in column H")
    print("Workbook saved with formulas preserved.")

if __name__ == "__main__":
    import_path = Path("C:/Users/Shadow/projects/analyitics_automator")
    update_sales_report(import_path)