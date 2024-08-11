# process_estimate_summary.py
# needs to process the 2023 estimate summary

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import os

def find_latest_sales_update(directory):
    sales_update_files = list(directory.glob("*Sales Update*.xlsx"))
    if not sales_update_files:
        raise FileNotFoundError("No Sales Update file found.")
    return max(sales_update_files, key=os.path.getmtime)

def process_estimate_summary(import_path):
    # Find the EstimateSummary file
    estimate_files = list(import_path.glob("*EstimateSummary*.csv"))
    if not estimate_files:
        print("No EstimateSummary file found.")
        return
    
    estimate_file = estimate_files[0]  # Use the first matching file
    print(f"Processing file: {estimate_file.name}")

    # Read the CSV file
    estimate_df = pd.read_csv(estimate_file, usecols=[
        'CustomerNumber', 'CustomerSize', 'ResidentialOrCommercial',
        'BranchNumberOfCustomer', 'ProgramCode', 'TotalPriceFormatted',
        'EstimateRequestedDate', 'EstimateGivenDate', 'RejectDate',
        'SoldDate', 'CancelDate'
    ])

    # Clean and convert TotalPriceFormatted
    estimate_df['TotalPriceFormatted'] = estimate_df['TotalPriceFormatted'].str.replace('$', '')
    estimate_df['TotalPriceFormatted'] = estimate_df['TotalPriceFormatted'].str.replace(',', '')
    estimate_df['TotalPriceFormatted'] = pd.to_numeric(estimate_df['TotalPriceFormatted'], errors='coerce')
    
    # Filter rows where TotalPriceFormatted > 0
    estimate_df = estimate_df[estimate_df['TotalPriceFormatted'] > 0].dropna(subset=['TotalPriceFormatted'])

    # Convert date columns to datetime
    date_columns = ['EstimateRequestedDate', 'EstimateGivenDate', 'RejectDate', 'SoldDate', 'CancelDate']
    for col in date_columns:
        estimate_df[col] = pd.to_datetime(estimate_df[col], errors='coerce')

    # Find the latest Sales Update file
    excel_file_path = find_latest_sales_update(import_path)
    print(f"Updating file: {excel_file_path.name}")

    # Load the existing Excel file
    book = load_workbook(excel_file_path, data_only=False)
    
    # Select the '2024 Estimate Summary' sheet
    sheet_name = '2024 Estimate Summary'
    if sheet_name not in book.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Please ensure the sheet exists.")
        return
    sheet = book[sheet_name]
    
    # Preserve formulas in columns F, M, N, O, P
    preserved_columns = {
        'F': [sheet.cell(row=i, column=6).value for i in range(1, sheet.max_row + 1)],
        'M': [sheet.cell(row=i, column=13).value for i in range(1, sheet.max_row + 1)],
        'N': [sheet.cell(row=i, column=14).value for i in range(1, sheet.max_row + 1)],
        'O': [sheet.cell(row=i, column=15).value for i in range(1, sheet.max_row + 1)],
        'P': [sheet.cell(row=i, column=16).value for i in range(1, sheet.max_row + 1)]
    }
    
    # Clear existing data in the sheet, except the preserved columns
    for row in sheet.iter_rows(min_row=2):  # Start from second row to preserve header
        for cell in row:
            if cell.column_letter not in preserved_columns:
                cell.value = None

    # Prepare column mapping
    columns = ['CustomerNumber', 'CustomerSize', 'ResidentialOrCommercial',
               'BranchNumberOfCustomer', 'ProgramCode', 'TotalPriceFormatted',
               'EstimateRequestedDate', 'EstimateGivenDate', 'RejectDate',
               'SoldDate', 'CancelDate']
    excel_columns = [1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 12]  # Skipping column F (6)

    # Write the header
    for col, header in zip(excel_columns, columns):
        sheet.cell(row=1, column=col, value=header if header != 'TotalPriceFormatted' else 'TotalPrice')

    # Write the data
    for row_idx, row in enumerate(estimate_df.itertuples(index=False), start=2):
        for col_idx, excel_col in enumerate(excel_columns):
            value = row[col_idx]
            cell = sheet.cell(row=row_idx, column=excel_col, value=value)
            
            if columns[col_idx] == 'TotalPriceFormatted':
                # Format TotalPrice as currency
                cell.number_format = '"$"#,##0.00'

    # Restore preserved columns
    for col, values in preserved_columns.items():
        for row, value in enumerate(values, start=1):
            sheet.cell(row=row, column=ord(col) - 64, value=value)

    # Calculate and save
    book.calculation.calcMode = 'auto'
    book.save(excel_file_path)


    print(f"Estimate summary data has been written to {excel_file_path.name} in the '{sheet_name}' tab")
    print(f"Total rows processed: {len(estimate_df)}")

if __name__ == "__main__":
    import_path = Path("C:/Users/Shadow/projects/analyitics_automator")
    process_estimate_summary(import_path)